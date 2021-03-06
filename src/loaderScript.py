import yaml
import json
import sys
import os.path
from openpyxl import load_workbook
import re
import requests
import unicodedata
from collections import Counter

# assumptions about the xlsx structure:
# [Concepts][        site name 1       ][        site name 2       ]...
# [        ][codesys][code][designation][codesys][code][designation]...
# ....rows....
#
# the first column of each "site name X" can be anything and have to be specified
# in the YAML config file
#
# refer to the documentation for more details

baseurl = 'http://localhost:5000/c3-cloud/'  # tests
# baseurl = 'http://cispro.chu-rouen.fr/c3-cloud/'


# colored terminal output
class bcolors:
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'


# helper function: print with a specified color
def printcolor(s, color, end='\n'):
    print(color, end='')
    print(s, end='')
    print(bcolors.ENDC, end=end)


# constants
SITE = 'site'
CODE_SYSTEM = 'code_system'
CODE_SYSTEM_URI = 'code_system_uri'
CODE_SYSTEM_VERSION = 'code_system_version'
CODE = 'code'
DESIGNATION = 'designation'
CONCEPT = 'concept'
EQUIVALENCE = 'equivalence'
EQUIVALENCE_DEFINITION = 'equivalence_definition'


# ### requests ### #

# normalize and trim a string
def clean(s):
    return(unicodedata.normalize('NFKC', s).strip())


# send a request to the REST API
# url: the part of the URL coming *after* the baseurl
# method: "get" or "post"
# data: json encoded data to send as the POST request body
# get: GET parameters
def sendrequest(url, method="get", get=None, data=None):
    full_url = "{}{}/".format(baseurl, url)
    args = {k: v for k, v in {'params': get, 'json': data}.items() if v}
    # execute either requests.get or requests.post based on the <method> arg
    r = getattr(requests, method)(url=full_url, **args)
    if r.status_code != 200:
        report['errors'] += 1
        print("error", r.status_code, r.text)
    return(r.text)


# send a mapping to upload
# o: python dict containing the mapping data
def upload_mapping(o):
    printcolor('[uploading <{}>@<{}>]'.format(o[CONCEPT], o[SITE]),
               color=bcolors.OKBLUE, end='')
    if len(o['codes']) > 0:
        try:
            # lookup the code system uri from the server
            uri = [cs[CODE_SYSTEM_URI] for cs in codesystems if cs[CODE_SYSTEM] == o['codes'][0][CODE_SYSTEM]][0]
        except IndexError:
            # if it doesn't exist, we generate a fake one
            # NB a better solution would be to provide a real one instead
            # not easy because of the difficulty to find oids
            uri = "uri:oid:{}_oid".format(o['codes'][0][CODE_SYSTEM])

        # add the uri to the POST data to send
        def f(c):
            c[CODE_SYSTEM_URI] = uri
            return(c)
        o['codes'] = [f(c) for c in o['codes']]
    sendrequest("mappings", method="post", data=o)


# ### model ### #

# given the dictionary of {concept: [mapping]},
# build the object to send as json data,
# check wether it already exist and decide to upload it or not
def process_items(d):
    # sorted lists for objects comparison
    def f(l):
        return(sorted([[v for k, v in x.items() if k in [CODE, DESIGNATION]]
                      for x in l]))

    # "pretty print"
    def disp(title, l):
        print('\n' +
              '\n'.join([title+':'] + [str([e, i[e]]) for i in l for e in sorted(i.keys())]) +
              '\n')

    for concept, sites in d.items():
        for site, items in sites.items():
            # items = [{clean(k): clean(v) for k, v in i.items()} for i in items]
            o = {
                'concept': clean(concept),
                'site': clean(site),
                'codes': [{clean(i): clean(item[i]) for i in [CODE_SYSTEM, CODE, DESIGNATION]}
                          for item in items]}

            # get the mappings that already exist on the server
            codelist = [i for i in mappings if i[CONCEPT] == o['concept']
                        and i[SITE] == o['site']]
            if len(codelist) > 0:
                print("already in db→ <{}>@<{}>:".format(concept, site), end='')
                if f(o['codes']) == f(codelist):
                    printcolor("[identical]", color=bcolors.OKGREEN, end='')
                    report['identical'] += 1
                else:
                    printcolor("[different]", color=bcolors.WARNING, end='')
                    disp('local', items)
                    disp('server', codelist)
                    report['different'] += 1

                    if(FORCE):
                        upload_mapping(o)
                    else:
                        printcolor('[use --force to overwrite]', color=bcolors.FAIL, end='')
            else:
                upload_mapping(o)
                print([e[CODE] for e in items], end='')
                report['new'] += 1
            print('')


# ### xlsx ### #


# coordinate manipulation
class Coord:
    def __init__(self, col, row):
        self.col = col
        self.row = row

    def __str__(self):
        return self.col + str(self.row)

    def __repr__(self):
        return self.__str__()

    def horizontal(self, dir=1):
        return Coord(chr(ord(self.col) + dir), self.row)

    def right(self, n=1):
        return self.horizontal(n)

    def left(self, n=1):
        return self.horizontal(-n)

    def top(coord):  # get the associated site
        return Coord(coord.col, 1)

    def concept(coord):  # get the associated concept
        return Coord("A", coord.row)


# retrieve a reference to another cell
# ie if the value of the cell is "=F42", retrieve the value at F42 in the current sheet
# and if it is "=$'some other sheet'.=F42", retrieve the F42 value in the sheet "some other sheet" in the workbook
def resolveCell(wb, sheet, coord):
    v = sheet[coord].value
    if isinstance(v, str) and v[0] == '=':
        if '.' in v:
            sht, coord = v.split('.')
            sht = wb[re.sub(r"^\$'|'$", "", sht)]
        else:
            sht = sheet
            coord = v
        coord = coord[1:]
        return(sht[coord].value)
    else:
        return(v)


def build_item(wb, sheet, startcoord):
    # print("building item for {}".format(startcoord),
    #       sheet[str(startcoord)].value)
    coords = {
        SITE: startcoord.top(),
        CODE_SYSTEM: startcoord,
        CODE:        startcoord.right(),
        DESIGNATION: startcoord.right(2),
        CONCEPT:     startcoord.concept()}

    return({k: str(resolveCell(wb, sheet, str(v)))
            for k, v in coords.items()})


def item_is_valid(i):
    return(not (i[CONCEPT] == "None"
           or i[CODE] == "None"))


# if the key is already present, append the item to d[key],
# otherwise create d[key]
def insert_item(d, item):
    try:
        d[item[CONCEPT]].append(item)
    except KeyError:
        d[item[CONCEPT]] = [item]


def group_by(l, key):
    keys = set([x[key] for x in l])
    return({x: [y for y in l if y[key] == x] for x in keys})


def importFile(f, config):
    d = {}
    wb = load_workbook(f)
    for sheetName in config['sheets']:
        sheet = wb[sheetName]
        for c in [a.upper() for a in config['columns']]:
            for r in range(3, sheet.max_row):
                # print(c, r)
                coord = Coord(c, r)
                i = build_item(wb, sheet, startcoord=coord)
                if item_is_valid(i):
                    # print(">>>>", i)
                    insert_item(d, i)
    d = {k: group_by(v, SITE) for k, v in d.items()}
    process_items(d)


# ### Main ### #


if __name__ == "__main__":
    jsondata = json.loads(sendrequest(url="all"))['data']
    mappings = jsondata['mappings']
    codesystems = jsondata['code_systems']

    report = Counter(
        identical=0,
        different=0,
        new=0,
        error=0
    )

    args = sys.argv

    FORCE = '--force' in args
    if FORCE:
        args.remove('--force')

    if len(args) < 2:
        raise(Exception("please provide the path to a yaml configuration file"))
    else:
        configpath = args[1]
        dirname = os.path.dirname(configpath)
        with open(configpath) as f:
            y = yaml.load(f)
            [importFile(os.path.join(dirname, k), e)
             for k, e in y.items()]

    print('done.')
    print('──────────────────────────────')
    for category, count in report.items():
        print('{}: {}'.format(category, count))
